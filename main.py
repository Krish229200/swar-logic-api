
import datetime
import io
import uuid
from pathlib import Path
from tempfile import NamedTemporaryFile

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from rapidfuzz import fuzz

app = FastAPI(
    title="ASR Accuracy Report API",
    description="Upload ground-truth vs ASR transcripts, get an Excel accuracy report back.",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
def fuzzy_align(gt_text: str, asr_text: str, threshold: float = 80.0):
    gt_words = gt_text.split()
    asr_remaining = asr_text.split()

    exact_words = []
    fuzzy_words = []

    for word in gt_words:
        if not asr_remaining:
            break
        if word in asr_remaining:
            exact_words.append(word)
            asr_remaining.remove(word)
            continue
        best_score, best_idx = -1.0, -1
        for idx, candidate in enumerate(asr_remaining):
            score = fuzz.ratio(word, candidate)
            if score > best_score:
                best_score, best_idx = score, idx

        if best_idx != -1 and best_score >= threshold:
            fuzzy_words.append(word)
            del asr_remaining[best_idx]

    return exact_words, fuzzy_words


def build_detail_report(
    df: pd.DataFrame,
    gt_col: str,
    asr_col: str,
    filename_col: str,
    partner_name: str,
    uploaded_by: str,
    threshold: float = 80.0,
) -> pd.DataFrame:
    uploaded_id = f"UID_{uuid.uuid4().hex[:8]}"
    uploaded_on = datetime.datetime.now()

    rows = []
    for i, row in df.iterrows():
        gt_raw = row.get(gt_col, "")
        asr_raw = row.get(asr_col, "")
        gt_text = "" if pd.isna(gt_raw) else str(gt_raw).strip().lower()
        asr_text = "" if pd.isna(asr_raw) else str(asr_raw).strip().lower()

        exact_words, fuzzy_words = fuzzy_align(gt_text, asr_text, threshold)

        gt_tokens = len(gt_text.split())
        exact_count = len(exact_words)
        fuzzy_count = len(fuzzy_words)
        matched_count = exact_count + fuzzy_count
        matched_pct = round(matched_count / gt_tokens * 100, 2) if gt_tokens else 0.0

        rows.append(
            {
                "id": 10000 + i,
                "Partner_Name": partner_name,
                "Filename": row.get(filename_col, ""),
                "GT_Translit": gt_raw if not pd.isna(gt_raw) else "",
                "ASR_Translit": asr_raw if not pd.isna(asr_raw) else "",
                "GT_Tokens": gt_tokens,
                "Exact_Words": ", ".join(exact_words) if exact_words else None,
                "Fuzzy_Words": ", ".join(fuzzy_words) if fuzzy_words else None,
                "Exact_Count": exact_count,
                "Fuzzy_Count": fuzzy_count,
                "Matched_Count": matched_count,
                "Matched_Ground_Truth": matched_pct,
                "Uploaded_id": uploaded_id,
                "uploaded_on": uploaded_on,
                "uploaded_by": uploaded_by,
            }
        )

    return pd.DataFrame(rows)


def build_summary_report(detail_df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        detail_df.groupby("Filename", sort=False)
        .agg(**{
            "Sum of GT_Tokens": ("GT_Tokens", "sum"),
            "Sum of Matched_Count": ("Matched_Count", "sum"),
        })
        .reset_index()
        .rename(columns={"Filename": "Row Labels"})
    )
    summary["ASR Accuracy %"] = (
        summary["Sum of Matched_Count"] / summary["Sum of GT_Tokens"]
    )

    total_gt = detail_df["GT_Tokens"].sum()
    total_matched = detail_df["Matched_Count"].sum()
    grand_total = pd.DataFrame(
        [
            {
                "Row Labels": "Grand Total",
                "Sum of GT_Tokens": total_gt,
                "Sum of Matched_Count": total_matched,
                "ASR Accuracy %": (total_matched / total_gt) if total_gt else 0.0,
            }
        ]
    )
    return pd.concat([summary, grand_total], ignore_index=True)


def write_report(detail_df: pd.DataFrame, summary_df: pd.DataFrame, output_path: str):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(
            writer, sheet_name="Accuracy Summary report", index=False, startrow=2
        )
        detail_df.to_excel(writer, sheet_name="Swar Data Report", index=False)
    wb = load_workbook(output_path)
    ws = wb["Accuracy Summary report"]
    header_row = 3
    acc_col = None
    for cell in ws[header_row]:
        if cell.value == "ASR Accuracy %":
            acc_col = cell.column_letter
            break
    if acc_col:
        for r in range(header_row + 1, ws.max_row + 1):
            ws[f"{acc_col}{r}"].number_format = "0.00%"
    wb.save(output_path)


def read_uploaded_table(upload: UploadFile, raw: bytes) -> pd.DataFrame:
    """Load an uploaded CSV or Excel file into a DataFrame."""
    name = (upload.filename or "").lower()
    buf = io.BytesIO(raw)
    try:
        if name.endswith(".csv"):
            return pd.read_csv(buf)
        elif name.endswith((".xlsx", ".xls")):
            return pd.read_excel(buf)
        else:
            # Fall back: sniff by trying Excel then CSV.
            try:
                buf.seek(0)
                return pd.read_excel(buf)
            except Exception:
                buf.seek(0)
                return pd.read_csv(buf)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse uploaded file: {exc}")

@app.get("/")
def health():
    return {"status": "ok", "docs": "/docs"}


@app.post("/report/columns")
async def preview_columns(file: UploadFile = File(...)):
    """
    Upload a CSV/XLSX first to see its column names, so you know what to pass
    as gt_col / asr_col / filename_col in /report/generate.
    """
    raw = await file.read()
    df = read_uploaded_table(file, raw)
    return {"columns": list(df.columns), "row_count": len(df)}


@app.post("/report/generate")
async def generate_report(
    file: UploadFile = File(..., description="CSV or XLSX with ground-truth and ASR transcript columns"),
    gt_col: str = Form("Actual transcript- G T", description="Column name for ground-truth transcript"),
    asr_col: str = Form("transcript", description="Column name for ASR transcript"),
    filename_col: str = Form("conversation_id", description="Column name identifying each conversation/file"),
    partner_name: str = Form(..., description="Partner name to stamp on every row"),
    uploaded_by: str = Form("Super_admin"),
    threshold: float = Form(80.0, description="Fuzzy match threshold (0-100)"),
):
    raw = await file.read()
    df = read_uploaded_table(file, raw)

    for col in (gt_col, asr_col, filename_col):
        if col not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Column '{col}' not found. Available columns: {list(df.columns)}",
            )

    detail_df = build_detail_report(
        df,
        gt_col=gt_col,
        asr_col=asr_col,
        filename_col=filename_col,
        partner_name=partner_name,
        uploaded_by=uploaded_by,
        threshold=threshold,
    )
    summary_df = build_summary_report(detail_df)

    # Write to a temp file so we can stream it back as a download.
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    write_report(detail_df, summary_df, tmp.name)

    base = Path(file.filename).stem if file.filename else "report"
    out_name = f"SwarReport_{partner_name}_{base}.xlsx"

    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=out_name,
    )
